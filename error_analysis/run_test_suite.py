import subprocess
import sys
import os
import pandas as pd
import json
import argparse
from openpyxl import load_workbook
from dotenv import load_dotenv
from llm import LLMProofWriter

load_dotenv()

"""
Runs the 18 test cases we have
"""


def test_parser():
    args = {
        "_on_rd_init_1": ["_on_rd_init", "_on_rd_init_precon_1"],
        "_on_rd_init_2": ["_on_rd_init", "_on_rd_init_precon_2"],
        "_on_rd_init_3": ["_on_rd_init", "_on_rd_init_precon_3"],
        "_on_rd_init_4": ["_on_rd_init", "_on_rd_init_precon_4"],
        "gcoap_dns_server_proxy_get_1": ["gcoap_dns_server_proxy_get", "gcoap_dns_server_proxy_get_precon_1"],
        "gcoap_dns_server_proxy_get_2": ["gcoap_dns_server_proxy_get", "gcoap_dns_server_proxy_get_precon_2"],
        "_gcoap_forward_proxy_copy_options_1": ["_gcoap_forward_proxy_copy_options", "_gcoap_forward_proxy_copy_options_precon_1"],
        "_gcoap_forward_proxy_copy_options_2": ["_gcoap_forward_proxy_copy_options", "_gcoap_forward_proxy_copy_options_precon_2"],
        "_iphc_ipv6_encode_1": ["_iphc_ipv6_encode", "_iphc_ipv6_encode_precon_1"],
        "_iphc_ipv6_encode_2": ["_iphc_ipv6_encode", "_iphc_ipv6_encode_precon_2"],
        "dns_msg_parse_reply_1": ["dns_msg_parse_reply", "dns_msg_parse_reply_precon_1"],
        "dns_msg_parse_reply_2": ["dns_msg_parse_reply", "dns_msg_parse_reply_precon_2"],
        "_rbuf_add_1": ["_rbuf_add2", "_rbuf_add_precon_1"],
        "_rbuf_add_2": ["_rbuf_add2", "_rbuf_add_precon_2"],
        "_rbuf_add_3": ["_rbuf_add2", "_rbuf_add_precon_3"],
        "_rbuf_add_4": ["_rbuf_add2", "_rbuf_add_precon_4"],
        "_rbuf_add_5": ["_rbuf_add2", "_rbuf_add_precon_5"],
        "_rbuf_add_6": ["_rbuf_add2", "_rbuf_add_precon_6"]
    }

    for case, args in args.items():
        print(f"\n===== Running {case} =====")
        result = subprocess.run([sys.executable, f"main.py"] + args, cwd=os.getcwd(), check=False)
        
        if result.returncode != 0:
            print(f"Error running {case}, return code: {result.returncode}")
            # Uncomment next line if you want to stop on first error
            # sys.exit(result.returncode)
        else:
            print(f"===== Completed {case} Successfully =====\n")



def test_workflow(tags=[], short_run=False):

    preconditions = {
        '_on_rd_init_precon_1': "__CPROVER_assume(hdr != NULL);",
        '_on_rd_init_precon_2': "__CPROVER_assume(payload_offset <= pkt_size);",
        '_on_rd_init_precon_3': "__CPROVER_assume(pkt != NULL);",
        '_on_rd_init_precon_4': "__CPROVER_assume(_result_buf != NULL);",
        'gcoap_dns_server_proxy_get_precon_1': "__CPROVER_assume(str != NULL);",
        'gcoap_dns_server_proxy_get_precon_2': "__CPROVER_assume(src_null_byte < CONFIG_GCOAP_DNS_SERVER_URI_LEN);",
        '_gcoap_forward_proxy_copy_options_precon_1': "__CPROVER_assume(client_pkt->payload_len <= pkt->payload_len - 1);",
        '_gcoap_forward_proxy_copy_options_precon_2': "__CPROVER_assume(size <= pkt->payload_len);",
        '_iphc_ipv6_encode_precon_1': "__CPROVER_assume(len >= 41);",
        '_iphc_ipv6_encode_precon_2': "__CPROVER_assume(data != NULL);",
        'dns_msg_parse_reply_precon_1': "__CPROVER_assume(family == AF_UNSPEC || family == AF_INET || family == AF_INET6);",
        'dns_msg_parse_reply_precon_2': "__CPROVER_assume(len >= sizeof(dns_hdr_t));",
        '_rbuf_add_precon_1': "__CPROVER_assume(offset != 0 || sixlowpan_frag_1_is(pkt->data) || sixlowpan_sfr_rfrag_is(data));",
        '_rbuf_add_precon_2': "__CPROVER_assume(offset_diff >= 0);",
        '_rbuf_add_precon_3': "__CPROVER_assume(datagram_size <= entry_size);",
        '_rbuf_add_precon_4': "__CPROVER_assume(offset < 1000);",
        '_rbuf_add_precon_5': "__CPROVER_assume(size > MAX(sizeof(sixlowpan_frag_t), MAX(sizeof(sixlowpan_frag_n_t), sizeof(sixlowpan_sfr_rfrag_t))));",
        '_rbuf_add_precon_6': "__CPROVER_assume(entry.pkt->data != NULL);"
    }

    fields = [
        'Tag',
        'Removed Precondition',
        'Testing Round',
        'Total Tokens Used',
        'Success',
        'Succeeded on Attempt',
        'Attempt #1 Response',
        'A#1 Input Tokens',
        'A#1 Output Tokens',
        'Attempt #2 Response',
        'A#2 Input Tokens',
        'A#2 Cached Tokens',
        'A#2 Output Tokens',
        'Attempt #3 Response',
        'A#3 Input Tokens',
        'A#3 Cached Tokens',
        'A#3 Output Tokens',
    ]

    summary = {
        'Attempt 1': 0,
        'Attempt 2': 0,
        'Attempt 3': 0,
        'Failed': 0
    }

    NUM_ROUNDS = 1
    openai_api_key = os.getenv("OPENAI_API_KEY", None)
    if not openai_api_key:
        raise EnvironmentError("No OpenAI API key found")

    results_file = os.path.join('./results', 'test_results.xlsx')
    raw_llm_responses = dict()

    df = pd.DataFrame(columns=fields)


    
    for tag, precon in preconditions.items():
        if len(tags) > 0 and tag not in tags or (short_run and (tag.startswith('_rbuf_add') or tag.startswith('_iphc_ipv6_encode'))):
            continue

        print(f"\n===== Running test for tag: {tag} =====")
        results = {
            'Tag': tag,
            'Removed Precondition': precon
        }

        for round_num in range(1, NUM_ROUNDS + 1):
            results["Testing Round"] = round_num
            try:
                total_tokens = 0
                proof_writer = LLMProofWriter(openai_api_key, tag, test_mode=True)
                attempts, llm_responses = proof_writer.iterate_proof(max_attempts=3)
                raw_llm_responses[tag] = {f'Attempt #{i + 1}': response['response'] for i, response in enumerate(llm_responses)}
                if attempts == 0:
                    results['Success'] = 'False'
                    summary['Failed'] += 1
                elif attempts == -1:
                    results['Success'] = 'Error'
                else:
                    results['Success'] = 'True'
                if results['Success'] == 'True':
                    results['Succeeded on Attempt'] = len(llm_responses)
                    summary[f'Attempt {len(llm_responses)}'] += 1


                for i, response in enumerate(llm_responses):
                    results[f'Attempt #{i + 1} Response'] = '\n '.join([f'{precon['precondition_as_code']} in {precon['function']}' for precon in response['response']['new_preconditions']])
                    results[f'A#{i + 1} Input Tokens'] = response['usage'].input_tokens
                    results[f'A#{i + 1} Output Tokens'] = response['usage'].output_tokens
                    if i > 0:
                        results[f'A#{i + 1} Cached Tokens'] = response['usage'].input_tokens_details.cached_tokens
                    total_tokens += response['usage'].total_tokens
                results['Total Tokens Used'] = total_tokens
            except Exception as e:
                print(f"Error during processing for tag {tag}: {e}")
                results['Success'] = 'Error'

            df = pd.concat([df, pd.DataFrame([results])])

    with pd.ExcelWriter(results_file, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name='Results', index=False)

    # Add the summary
    workbook = load_workbook(results_file)
    worksheet = workbook['Results']
    startrow = worksheet.max_row + 2  # +2 for spacing
    summary_df = pd.DataFrame(
        list(summary.items()), columns=['Succeeded On', 'Count']
    )
    with pd.ExcelWriter(results_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            summary_df.to_excel(writer, sheet_name='Results', index=False, startrow=startrow)

    with open('./results/response_dump.json', 'w') as f:
        json.dump(raw_llm_responses, f, indent=4)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Test runner configuration")
    parser.add_argument(
        "--tags",
        nargs="*",
        default=[],
        help="List of harness git tags to use in the test run. Invalid tags will be skipped. If no tags are provided, all test tags will be run.",
    )
    parser.add_argument(
        "--parser_only",
        action="store_true",
        help="Only run the parser component of the test suite"
    )
    parser.add_argument(
        "--short",
        action="store_true",
        help="Skip tests for _rbuf_add and _iphc_ipv6_encode due to long runtime"
    )
    args = parser.parse_args()

    if args.parser_only:
        test_parser()

    else:
        test_workflow(tags=args.tags, short_run=args.short)